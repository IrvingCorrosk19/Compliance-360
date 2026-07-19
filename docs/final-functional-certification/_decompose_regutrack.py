"""Decompose REGUTRACK Excel into JSON inventory for certification."""
from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Proyectos\Compliance 360\REGUTRACK 02JUN26 MG.xlsx")
OUT = Path(r"c:\Proyectos\Compliance 360\docs\final-functional-certification\evidence\regutrack_decomposition.json")

wb = openpyxl.load_workbook(XLSX, data_only=False, read_only=True)
result = {"file": str(XLSX.name), "sheets": []}

for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    headers = []
    header_row_idx = 0
    if rows:
        # find first non-empty header-like row
        for i, row in enumerate(rows[:5]):
            non_empty = [c for c in row if c is not None and str(c).strip() != ""]
            if len(non_empty) >= 3:
                headers = [str(c).strip() if c is not None else f"COL_{j+1}" for j, c in enumerate(row)]
                header_row_idx = i
                break
    data_rows = rows[header_row_idx + 1 :] if rows else []
    filled = 0
    for r in data_rows:
        if any(c is not None and str(c).strip() != "" for c in (r or [])):
            filled += 1

    col_stats = []
    for ci, h in enumerate(headers):
        values = []
        for r in data_rows:
            if r is None or ci >= len(r):
                continue
            v = r[ci]
            if v is None or str(v).strip() == "":
                continue
            values.append(v)
        sample = [str(v)[:80] for v in values[:5]]
        unique = len({str(v) for v in values})
        col_stats.append(
            {
                "index": ci + 1,
                "header": h,
                "nonEmpty": len(values),
                "uniqueApprox": unique,
                "samples": sample,
            }
        )

    result["sheets"].append(
        {
            "name": name,
            "maxRow": ws.max_row,
            "maxCol": ws.max_column,
            "headerRow": header_row_idx + 1,
            "headerCount": len([h for h in headers if h and not h.startswith("COL_")]),
            "dataRowsFilled": filled,
            "columns": col_stats,
        }
    )

wb.close()
OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
print("sheets=", len(result["sheets"]))
for s in result["sheets"]:
    print(f"- {s['name']}: cols={s['headerCount']} rows={s['dataRowsFilled']} max={s['maxRow']}x{s['maxCol']}")
print("wrote", OUT)
