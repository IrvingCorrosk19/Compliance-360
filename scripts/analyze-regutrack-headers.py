"""Extract full headers and column groups from REGUTRACK."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path(r"c:\Proyectos\Compliance 360\REGUTRACK 02JUN26 MG.xlsx")
wb = load_workbook(XLSX, data_only=True, read_only=True)

def headers_from(ws, header_row=1, max_cols=120):
    row = next(ws.iter_rows(min_row=header_row, max_row=header_row, max_col=max_cols, values_only=True))
    out = []
    for i, v in enumerate(row, start=1):
        if v is None or str(v).strip() == "":
            continue
        out.append({"col": i, "header": str(v).replace("\n", " / ").strip()})
    return out

def group_header(h: str) -> str:
    t = h.lower()
    if any(k in t for k in ["copia", "certificado", "declaraci", "etiqueta", "instructivo", "literatura", "iso", "manual", "ensayo", "estudio", "croquis", "timbre", "tasa", "aviso", "regente", "cedula", "cédula", "pasaporte", "registro publico", "contrato", "lista de dispositivos", "requisito", "documento"]):
        return "checklist_docs"
    if any(k in t for k in ["fecha", "vencim", "expir", "armado", "sometim", "aprobac", "recep", "solicitud"]):
        return "fechas_workflow"
    if any(k in t for k in ["oportunidad", "prioridad", "sales", "mkt", "comercial"]):
        return "comercial"
    if any(k in t for k in ["observa", "comentario", "seguimiento", "estatus", "status", "cuello", "aging", "detenid"]):
        return "seguimiento"
    if any(k in t for k in ["país", "pais", "categoria", "marca", "producto", "descrip", "catálogo", "catalogo", "fabricante", "distribuidor", "iniciativa", "ficha", "formulario", "entidad", "criterio", "registro sanitario", "clase", "tipo de proceso", "proveedor"]):
        return "identidad_producto"
    return "otros"

result = {}

# CTT REGISTROS (2)
ws = wb["CTT REGISTROS (2)"]
h = headers_from(ws, 1, 120)
groups = {}
for item in h:
    g = group_header(item["header"])
    groups.setdefault(g, []).append(item["header"])
result["CTT REGISTROS (2)"] = {
    "nonempty_approx": 192,
    "header_count": len(h),
    "all_headers": [x["header"] for x in h],
    "groups": {k: len(v) for k, v in groups.items()},
    "groups_detail": groups,
}

# DOCUMENTACION - row1 headers, also special cols
ws = wb["DOCUMENTACION"]
h = headers_from(ws, 1, 40)
result["DOCUMENTACION"] = {
    "header_count": len(h),
    "all_headers": [x["header"] for x in h],
}

# CTT LICENCIAS OP - company header rows 1-3, table header row 5
ws = wb["CTT LICENCIAS OP"]
# rows 1-3 company
companies = []
row1 = list(next(ws.iter_rows(min_row=1, max_row=1, max_col=15, values_only=True)))
row2 = list(next(ws.iter_rows(min_row=2, max_row=2, max_col=15, values_only=True)))
row3 = list(next(ws.iter_rows(min_row=3, max_row=3, max_col=15, values_only=True)))
for i, v in enumerate(row1):
    if v and str(v).strip():
        companies.append({
            "col": i + 1,
            "company": str(v).strip(),
            "row2": None if row2[i] is None else str(row2[i]),
            "row3": None if row3[i] is None else str(row3[i]),
        })
h5 = headers_from(ws, 5, 40)
# also check row4
h4 = headers_from(ws, 4, 40)
result["CTT LICENCIAS OP"] = {
    "company_block": companies,
    "row4_headers": [x["header"] for x in h4],
    "row5_headers": [x["header"] for x in h5],
}

# TUBERIA
ws = wb["CTT REGISTROS TUBERIA"]
h = headers_from(ws, 1, 100)
groups = {}
for item in h:
    g = group_header(item["header"])
    groups.setdefault(g, []).append(item["header"])
result["CTT REGISTROS TUBERIA"] = {
    "header_count": len(h),
    "all_headers": [x["header"] for x in h],
    "groups": {k: len(v) for k, v in groups.items()},
}

# Compare CTT REGISTROS vs (2)
a = set(result["CTT REGISTROS (2)"]["all_headers"])
ws = wb["CTT REGISTROS"]
h = headers_from(ws, 1, 120)
b = set(x["header"] for x in h)
result["CTT REGISTROS"] = {
    "header_count": len(h),
    "same_headers_as_2": a == b,
    "only_in_2": sorted(a - b)[:30],
    "only_in_base": sorted(b - a)[:30],
}

wb.close()
out = Path(r"c:\Proyectos\Compliance 360\docs\user-manual\docs\_regutrack_headers.json")
out.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
print("CTT REGISTROS (2) headers:", result["CTT REGISTROS (2)"]["header_count"], "groups", result["CTT REGISTROS (2)"]["groups"])
print("TUBERIA headers:", result["CTT REGISTROS TUBERIA"]["header_count"], "groups", result["CTT REGISTROS TUBERIA"]["groups"])
print("DOCUMENTACION:", result["DOCUMENTACION"]["all_headers"])
print("LICENCIAS companies:", result["CTT LICENCIAS OP"]["company_block"])
print("LICENCIAS row5:", result["CTT LICENCIAS OP"]["row5_headers"])
print("REGISTROS same as (2):", result["CTT REGISTROS"]["same_headers_as_2"])
print("\n--- checklist_docs sample ---")
for x in result["CTT REGISTROS (2)"]["groups_detail"].get("checklist_docs", [])[:25]:
    print(" -", x)
print("\n--- fechas ---")
for x in result["CTT REGISTROS (2)"]["groups_detail"].get("fechas_workflow", []):
    print(" -", x)
print("\n--- identidad ---")
for x in result["CTT REGISTROS (2)"]["groups_detail"].get("identidad_producto", []):
    print(" -", x)
print("\n--- seguimiento ---")
for x in result["CTT REGISTROS (2)"]["groups_detail"].get("seguimiento", []):
    print(" -", x)
print("\n--- comercial ---")
for x in result["CTT REGISTROS (2)"]["groups_detail"].get("comercial", []):
    print(" -", x)
print("\n--- otros ---")
for x in result["CTT REGISTROS (2)"]["groups_detail"].get("otros", []):
    print(" -", x)
