"""Deep sheet-by-sheet analysis of REGUTRACK workbook."""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path(r"c:\Proyectos\Compliance 360\REGUTRACK 02JUN26 MG.xlsx")
OUT = Path(r"c:\Proyectos\Compliance 360\docs\user-manual\docs\_regutrack_sheet_probe.json")

wb = load_workbook(XLSX, data_only=True, read_only=True)
report = {"file": str(XLSX), "sheets": []}

for name in wb.sheetnames:
    ws = wb[name]
    # Collect first ~8 rows to detect headers (REGUTRACK often has multi-row headers)
    preview_rows = []
    max_col = 0
    row_count = 0
    nonempty_rows = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_count = i
        vals = list(row)
        # trim trailing Nones
        while vals and vals[-1] is None:
            vals.pop()
        if any(v is not None and str(v).strip() != "" for v in vals):
            nonempty_rows += 1
            max_col = max(max_col, len(vals))
            if len(preview_rows) < 12:
                preview_rows.append([None if v is None else (str(v)[:120] if not isinstance(v, (int, float)) else v) for v in vals[:80]])
        if i >= 8000:  # safety
            break

    # Heuristic header: first nonempty row with most text cells
    header_candidates = []
    for idx, r in enumerate(preview_rows[:6], start=1):
        textish = sum(1 for v in r if isinstance(v, str) and len(v.strip()) > 1)
        header_candidates.append({"row": idx, "text_cells": textish, "sample": r[:25]})

    report["sheets"].append({
        "name": name,
        "total_rows_scanned": row_count,
        "nonempty_rows": nonempty_rows,
        "approx_cols": max_col,
        "header_candidates": header_candidates,
        "preview_rows": preview_rows[:8],
    })

wb.close()
OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
print(json.dumps({"sheet_count": len(report["sheets"]), "names": [s["name"] for s in report["sheets"]]}, ensure_ascii=False, indent=2))
for s in report["sheets"]:
    print(f"\n=== {s['name']} | nonempty={s['nonempty_rows']} cols~{s['approx_cols']} ===")
    if s["header_candidates"]:
        best = max(s["header_candidates"], key=lambda x: x["text_cells"])
        print("HEADER-ish row", best["row"], ":", best["sample"][:20])
